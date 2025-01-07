import streamlit as st
import fitz  # PyMuPDF
from openpyxl import load_workbook
import requests
import tempfile

# PDFからテキストを抽出する関数
def extract_text_from_pdf(pdf_file, rects):
    text_list = []
    labels = []
    
    # Streamlitでアップロードされたファイルを一時ファイルとして保存
    with fitz.open(stream=pdf_file.read(), filetype="pdf") as doc:  
        for rect_data in rects:
            x0, y0, x1, y1, label = rect_data
            rect = fitz.Rect(x0, y0, x1, y1)
            labels.append(label)
            text = ""
            for page in doc:
                extracted_text = page.get_text("text", clip=rect)
                text += extracted_text.strip() if extracted_text.strip() else ""
            text_list.append(text)
    return text_list, labels

# StreamlitによるGUI
def main(uploaded_file):
    st.title("伝票作成アプリ")

    # PDFアップロード
    if uploaded_file:
        st.success(f"{uploaded_file.name} がアップロードされました")

        # 抽出範囲の指定
        rects = [
            (140, 175, 180, 190, 'AC9-1'),
            (140, 190, 180, 205, 'AC9'),
            (210, 190, 500, 205, 'AC11'),
            (140, 205, 180, 220, 'AC13'),
            (210, 205, 500, 220, 'AC15'),
            (100, 215, 140, 230, 'AC17'),
            (135, 220, 500, 230, 'AC19'),
            (105, 295, 250, 305, 'A11'),
            (400, 360, 500, 370, 'S11')
        ]

        # テキスト抽出
        text_list, labels = extract_text_from_pdf(uploaded_file, rects)
        st.write("抽出されたテキスト:")
        for label, text in zip(labels, text_list):
            st.write(f"**{label}**: {text}")

        # ユーザー入力
        syukka = st.date_input("出荷日を入力してください")
        buturyu = st.selectbox("物流センターを選択してください",['AX44', 'AX60', 'AX02','AX36','AX86','AX28'])
        konpou = st.selectbox("梱包数を選択してください",['1','2','3','4','5','それ以上'])
        if konpou =='それ以上':
            konpou = st.text_input('梱包数を入力してください')

        # Excel更新
        if st.button("Excelファイルを生成する"):
            # GitHubのリポジトリURL
            import shutil
            github_url = "https://github.com/shintarotakasaki/excel3/raw/main/伝票(規格品)_ラベル_指示書.xlsm"

            # ファイルをダウンロードして一時ファイルとして保存
            response = requests.get(github_url,stream=True)
            if response.status_code == 200:
                file_path = "伝票(規格品)_ラベル_指示書.xlsm"  # ここで file_path を定義
                with open("伝票(規格品)_ラベル_指示書.xlsm",'wb')as f:
                    response.raw.decode_content = True
                    shutil.copyfileobj(response.raw, f)
            
            try:
                wb = load_workbook(file_path, keep_vba=True)
                ws = wb['納品書控(製品)']
                wb.active = ws
                
                # Excelファイルへの書き込み
                ws['AH3'] = syukka
                ws['AM9'] = buturyu
                ws['AB4'] = konpou + "梱包"

                zig_tok = ""
                for i, text in enumerate(text_list):
                    if labels[i] == 'AC9-1':
                        zig_tok = text
                    elif labels[i] == 'AC9':
                        ws[labels[i]] = zig_tok + '-' + text
                    elif labels[i] == 'AC11':
                        ws[labels[i]] = text + "様"
                    elif labels[i] == 'AC13':
                        ws[labels[i]] = '届け先：' + text
                    elif labels[i] == 'AC15':
                        ws[labels[i]] = text + "様" if text else "=AC11"
                    elif labels[i] == 'AC17':
                        ws[labels[i]] = '現場名：' + text
                    else:
                        ws[labels[i]] = text

                # 保存とダウンロード
                wb.save(file_path)
                st.success("Excelファイルが上書き保存されました！")                

                #ファイルをダウンロード
                st.write("保存されたファイルを以下のリンクからダウンロードしてください:")
                with open(file_path, "rb") as file:
                    st.download_button(
                        label="ダウンロードする",
                        data=file,
                        file_name="伝票(規格品)_ラベル_指示書.xlsm",
                        mime="application/vnd.ms-excel"
                    )
            except Exception as e:
                st.error(f"エラーが発生しました: {e}")                

if __name__ == "__main__":
    main()
